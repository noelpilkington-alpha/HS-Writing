"""Build the source-first spreadsheet twin of the Primary Source Register.
Tab 1 = primary-first; Tab 2 = the original artifact-first view (preserved)."""
import os, re, csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COURSE = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
REG = os.path.join(COURSE, "BrainLift_Primary_Source_Register.md")
XLSX = os.path.join(COURSE, "BrainLift_Knowledge_Tree_Sources.xlsx")
CSV = os.path.join(COURSE, "BrainLift_Knowledge_Tree_Sources.csv")

def parse_register(md):
    """Yield (category, source, typ, grounds, via, compiled) from the register's category tables."""
    cat = None
    for line in md.splitlines():
        m = re.match(r'^##\s+Category\s+(.+)$', line)
        if m:
            cat = "Category " + m.group(1).strip()
            continue
        if line.startswith("## ") and not m:
            cat = None
            continue
        if line.startswith("|") and "---" not in line and cat:
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            if len(cells) == 5 and cells[1] in ("Authority", "Expert"):
                yield (cat, cells[0], cells[1], cells[2], cells[3], cells[4])

rows = list(parse_register(open(REG, encoding="utf-8").read()))

wb = Workbook()
ws = wb.active; ws.title = "Sources (primary-first)"
hdr = ["Primary Source", "Type", "Grounds (claim / SPOV)", "Reachable via",
       "Compiled into (artifacts)", "Knowledge Tree Category"]
ws.append(hdr)
for c in range(1, len(hdr)+1):
    cell = ws.cell(row=1, column=c); cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
for (cat, src, typ, grounds, via, comp) in rows:
    ws.append([src, typ, grounds, via, comp, cat])
widths = [46, 10, 52, 40, 46, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# preserve the original artifact-first sheet as a second tab, if present
if os.path.isfile(XLSX):
    old = load_workbook(XLSX)
    src_ws = old[old.sheetnames[0]]
    ws2 = wb.create_sheet("Artifacts (original)")
    for row in src_ws.iter_rows(values_only=True):
        ws2.append(row)

wb.save(XLSX)

with open(CSV, "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh); w.writerow(hdr)
    for (cat, src, typ, grounds, via, comp) in rows:
        w.writerow([src, typ, grounds, via, comp, cat])

print(f"source-first rows: {len(rows)}")
print(f"tabs: {wb.sheetnames}")
